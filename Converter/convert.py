from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class Vertex:
    def __init__(self, id, x, y):
        self.x = x
        self.y = y
        self.id = id

class Arc:
    def __init__(self, id, start, end):
        self.id = id
        self.start = start
        self.end = end

filename = "SETOR_01.xlsx"
wb = load_workbook(filename = filename)
ws = wb.active 

# iterate through the table rows, adding the points
coordsMap = {}
vertices = {}
count = -1
for row in ws.values:
    if count != -1:
        location = row[5]
        id = row[0]
        x = row[6]
        y = row[7]

        if id not in vertices.keys():
            vertices[id] = {}

        # we need this because points can have same coordinates...
        if (x, y) not in coordsMap.keys():
            v = Vertex(count, x, y)
            vertices[id][location] = v 
            coordsMap[(x, y)] = v
        else:
            v = coordsMap[(x, y)]
            vertices[id][location] = v
            count -= 1
        
    count += 1

# iterate through the points, creating the arcs
vertices2 = {}
arcs = {}
for id in vertices.keys():
    if "F" in vertices[id].keys():
        arcs[id] = Arc(id, vertices[id]["I"], vertices[id]["F"])
        vertices2[vertices[id]["I"].id] = vertices[id]["I"]
        vertices2[vertices[id]["F"].id] = vertices[id]["F"]
    else:
        arcs[id] = Arc(id, vertices[id]["I"], vertices[id]["M"])
        vertices2[vertices[id]["I"].id] = vertices[id]["I"]
        vertices2[vertices[id]["M"].id] = vertices[id]["M"]

# output a file in standard format
output = open("graph.in", "w")

output.write(str(len(vertices2)) + " " + str(len(arcs)) + "\n")
output.write("VERTICES:\n")
for i in vertices2.keys():
    output.write(str(i) + " " + str(vertices2[i].x) + " " + str(vertices2[i].y) + "\n")

output.write("ARCS:\n")
for i in arcs.keys():
    output.write(str(arcs[i].start.id) + " " + str(arcs[i].end.id) + "\n")